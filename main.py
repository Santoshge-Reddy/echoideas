from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait 
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.select import Select
from logging.handlers import TimedRotatingFileHandler
from selenium.webdriver.common.keys import Keys
from selenium.common.exceptions import NoSuchElementException , ElementClickInterceptedException, ElementNotInteractableException, TimeoutException
from bs4 import BeautifulSoup

import os
import time
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook




class ecoideaz():
	"""docstring for ecoideaz"""
	def __init__(self):
		super(ecoideaz, self).__init__()
		# self.arg = arg

		# Art & Crafts
		# self.url = 'https://www.ecoideaz.com/green-directory/business-category/alternative-energy-products-in-india/?w2dc_action=search&hash=e466dbfc00ec2e1ee7c3f50402ef55d7&controller=listings_controller&include_categories_children=1&categories=132&categories_path=Art+%26+Crafts+&what_search=&location_id=0&location_id_path=&address='
		# Alternative Energy
		# self.url = 'https://www.ecoideaz.com/green-directory/business-category/alternative-energy-products-in-india/?w2dc_action=search&hash=e466dbfc00ec2e1ee7c3f50402ef55d7&controller=listings_controller&include_categories_children=1&categories=131&categories_path=Alternative+Energy&what_search=&location_id=0&location_id_path=&address='
		# Organic Food & Drinks
		# self.url = 'https://www.ecoideaz.com/green-directory/business-category/alternative-energy-products-in-india/?w2dc_action=search&hash=e466dbfc00ec2e1ee7c3f50402ef55d7&controller=listings_controller&include_categories_children=1&categories=165&categories_path=Organic+Food+%26+Drinks+&what_search=&location_id=129&location_id_path=India&address='
		# green funding
		# self.url = 'https://www.ecoideaz.com/green-directory/business-category/alternative-energy-products-in-india/?w2dc_action=search&hash=e466dbfc00ec2e1ee7c3f50402ef55d7&controller=listings_controller&include_categories_children=1&categories=139&categories_path=Green+Funding+&what_search=&location_id=0&location_id_path=&address='
		# Beauty & Personal Care
		# self.url = 'https://www.ecoideaz.com/green-directory/business-category/alternative-energy-products-in-india/?w2dc_action=search&hash=e466dbfc00ec2e1ee7c3f50402ef55d7&controller=listings_controller&include_categories_children=1&categories=188&categories_path=Beauty+%26+Personal+Care+&what_search=&location_id=0&location_id_path=&address='
		

		# E-vehicles
		# self.url = 'https://www.ecoideaz.com/green-directory/business-category/alternative-energy-products-in-india/?w2dc_action=search&hash=e466dbfc00ec2e1ee7c3f50402ef55d7&controller=listings_controller&include_categories_children=1&categories=243&categories_path=E-vehicles+&what_search=&location_id=0&location_id_path=&address='
		# Eco Fashion
		# self.url = 'https://www.ecoideaz.com/green-directory/business-category/alternative-energy-products-in-india/?w2dc_action=search&hash=e466dbfc00ec2e1ee7c3f50402ef55d7&controller=listings_controller&include_categories_children=1&categories=134&categories_path=Eco+Fashion+&what_search=&location_id=0&location_id_path=&address='
		# Eco-friendly Packaging
		# self.url = 'https://www.ecoideaz.com/green-directory/business-category/alternative-energy-products-in-india/?w2dc_action=search&hash=e466dbfc00ec2e1ee7c3f50402ef55d7&controller=listings_controller&include_categories_children=1&categories=194&categories_path=Eco-friendly+Packaging+&what_search=&location_id=0&location_id_path=&address='
		# Eco-tourism
		# self.url = 'https://www.ecoideaz.com/green-directory/business-category/alternative-energy-products-in-india/?w2dc_action=search&hash=e466dbfc00ec2e1ee7c3f50402ef55d7&controller=listings_controller&include_categories_children=1&categories=141&categories_path=Eco-tourism+&what_search=&location_id=0&location_id_path=&address='
		# Green Architecture 
		# self.url = 'https://www.ecoideaz.com/green-directory/business-category/alternative-energy-products-in-india/?w2dc_action=search&hash=e466dbfc00ec2e1ee7c3f50402ef55d7&controller=listings_controller&include_categories_children=1&categories=133&categories_path=Green+Architecture+&what_search=&location_id=0&location_id_path=&address='
		# Green Gadgets 
		# self.url = 'https://www.ecoideaz.com/green-directory/business-category/alternative-energy-products-in-india/?w2dc_action=search&hash=e466dbfc00ec2e1ee7c3f50402ef55d7&controller=listings_controller&include_categories_children=1&categories=238&categories_path=Green+Gadgets+&what_search=&location_id=0&location_id_path=&address='
		# Green Innovations
		# self.url = 'https://www.ecoideaz.com/green-directory/business-category/alternative-energy-products-in-india/?w2dc_action=search&hash=e466dbfc00ec2e1ee7c3f50402ef55d7&controller=listings_controller&include_categories_children=1&categories=138&categories_path=Green+Innovations+&what_search=&location_id=0&location_id_path=&address='
		# Green Media
		# self.url = 'https://www.ecoideaz.com/green-directory/business-category/alternative-energy-products-in-india/?w2dc_action=search&hash=e466dbfc00ec2e1ee7c3f50402ef55d7&controller=listings_controller&include_categories_children=1&categories=137&categories_path=Green+Media+&what_search=&location_id=0&location_id_path=&address='
		# Green Products
		# self.url = 'https://www.ecoideaz.com/green-directory/business-category/alternative-energy-products-in-india/?w2dc_action=search&hash=e466dbfc00ec2e1ee7c3f50402ef55d7&controller=listings_controller&include_categories_children=1&categories=148&categories_path=Green+Products+&what_search=&location_id=0&location_id_path=&address='
		# Herbal medicine 
		# self.url = 'https://www.ecoideaz.com/green-directory/business-category/alternative-energy-products-in-india/?w2dc_action=search&hash=e466dbfc00ec2e1ee7c3f50402ef55d7&controller=listings_controller&include_categories_children=1&categories=240&categories_path=Herbal+medicine+&what_search=&location_id=0&location_id_path=&address='
		# Organic Agriculture
		# self.url = 'https://www.ecoideaz.com/green-directory/business-category/alternative-energy-products-in-india/?w2dc_action=search&hash=e466dbfc00ec2e1ee7c3f50402ef55d7&controller=listings_controller&include_categories_children=1&categories=143&categories_path=Organic+Agriculture+&what_search=&location_id=0&location_id_path=&address='
		# Rural Development 
		# self.url = 'https://www.ecoideaz.com/green-directory/business-category/alternative-energy-products-in-india/?w2dc_action=search&hash=e466dbfc00ec2e1ee7c3f50402ef55d7&controller=listings_controller&include_categories_children=1&categories=135&categories_path=Rural+Development+&what_search=&location_id=0&location_id_path=&address='
		# Training & Education 
		# self.url = 'https://www.ecoideaz.com/green-directory/business-category/alternative-energy-products-in-india/?w2dc_action=search&hash=e466dbfc00ec2e1ee7c3f50402ef55d7&controller=listings_controller&include_categories_children=1&categories=155&categories_path=Training+%26+Education+&what_search=&location_id=0&location_id_path=&address='
		# Waste Management
		# self.url = 'https://www.ecoideaz.com/green-directory/business-category/alternative-energy-products-in-india/?w2dc_action=search&hash=e466dbfc00ec2e1ee7c3f50402ef55d7&controller=listings_controller&include_categories_children=1&categories=136&categories_path=Waste+Management+&what_search=&location_id=0&location_id_path=&address='
		# Water Conservation 
		self.url = 'https://www.ecoideaz.com/green-directory/business-category/alternative-energy-products-in-india/?w2dc_action=search&hash=e466dbfc00ec2e1ee7c3f50402ef55d7&controller=listings_controller&include_categories_children=1&categories=157&categories_path=Water+Conservation+&what_search=&location_id=0&location_id_path=&address='
		

		self.information = {}
		self.scrap()

	def scrap(self):
		options = webdriver.ChromeOptions()
		options.add_argument("start-maximized")
		options.add_experimental_option("excludeSwitches", ["enable-automation"])
		options.add_experimental_option('useAutomationExtension', False)
		options.add_argument("--disable-notifications")
		# config = configparser.ConfigParser()
		prefs = {"plugins.always_open_pdf_externally": True}
		options.add_experimental_option("prefs",prefs)
		options.add_experimental_option("excludeSwitches",["ignore-certificate-errors"])
		# options.add_argument('--disable-gpu')
		# options.add_argument('--headless')
		dirpath = os.getcwd()
		foldername = os.path.dirname(os.path.realpath(__file__))
		driver = webdriver.Chrome(options=options, executable_path= foldername + '/chromedriver.exe')
		driver.get(self.url)
		self.open(driver)

	def open(self, driver):
		# driver.get(self.url)
		# print(driver.find_element_by_xpath('//*[@id="w2dc-controller-e466dbfc00ec2e1ee7c3f50402ef55d7"]/div/button'))
		nextButton = []
		try:
			nextButton = WebDriverWait(driver, 10).until(EC.presence_of_all_elements_located((By.XPATH, '//*[@id="w2dc-controller-e466dbfc00ec2e1ee7c3f50402ef55d7"]/div/button')))
			# nextButton[0].click()
		except TimeoutException:
			self.collectData(driver)
			# print(self.information)

		if len(nextButton) > 0:
			try:
				nextButton[0].click()
				self.open(driver)
			except ElementClickInterceptedException:
				# time.sleep(30)
				# driver.implicitly_wait(30)
				# if driver.find_element_by_xpath('//*[@id="custom_field_submit_102973"]'):
				# driver.switch_to_alert()
				try:
					driver.find_element_by_xpath('/html/body/div/a').click()
				except Exception as e:
					print(e)

				# print('waited 30')
				# prevButton = WebDriverWait(driver, 30).until(EC.presence_of_all_elements_located((By.XPATH, "/html/body/div/a")))
				# prevButton[0].click()
				# print('==========================')
				# print('clicked')
				# print('==========================')
				self.open(driver)

			except ElementNotInteractableException:
				self.collectData(driver)
			except TimeoutException:
				pass



	def collectData(self,driver):
		articles = WebDriverWait(driver, 10).until(EC.presence_of_all_elements_located((By.XPATH, '//article')))
		for article in articles:
			soup = BeautifulSoup(article.get_attribute('innerHTML'), 'lxml')
			title = soup.select_one('header a').text
			category = soup.select_one('.w2dc-label-primary a').text
			href = soup.select_one('header a')['href']

			self.information[title] = { 
                                       'title' : title,
                                       'category': category,
                                       'link': href
                                      }


		for info in self.information:
			# print(self.information[info])
			url = self.information[info]['link']
			driver.get(url)
			articles = WebDriverWait(driver, 10).until(EC.presence_of_all_elements_located((By.XPATH, '//article')))
			soup = BeautifulSoup(articles[0].get_attribute('innerHTML'), 'lxml')
			# description = soup.select_one('.w2dc-field-content.w2dc-field-description').text
			# website = soup.select_one('.w2dc-field-output-block-website a').text
			# address = soup.select_one('.w2dc-location .w2dc-show-on-map').text
			# email = soup.select_one('.w2dc-field-output-block.w2dc-field-output-block-email .w2dc-field-content a').text

			try:
				self.information[info]['description'] = BeautifulSoup(soup.select_one('.w2dc-field-content.w2dc-field-description').text, "lxml").text
			except Exception as e:
				pass

			try:
				self.information[info]['email'] = soup.select_one('.w2dc-field-output-block.w2dc-field-output-block-email .w2dc-field-content a').text
				
			except Exception as e:
				pass


			try:
				self.information[info]['address'] = BeautifulSoup(soup.select_one('.w2dc-location .w2dc-show-on-map').text, "lxml").text 
			except Exception as e:
				pass


			try:
				self.information[info]['website'] = soup.select_one('.w2dc-field-output-block-website a').text
			except Exception as e:
				pass

		sheet_name = self.information[info]['category']
		# print('to excel')
		df_information = pd.DataFrame(self.information)
		result = df_information.transpose() 
		# result.to_excel("output.xlsx")

		path = r"output.xlsx"
		# writer = pd.ExcelWriter(path, engine='xlsxwriter')

		writer = pd.ExcelWriter(path, engine='openpyxl')
		book = load_workbook(path)
		writer.book = book

		# if os.path.exists(path):
		#     book = openpyxl.load_workbook(file_name)
		#     writer.book = book

		result.to_excel(writer, sheet_name = sheet_name)
		writer.save()
		writer.close()



if __name__ == '__main__':
	ecoideaz()
	# print(result)